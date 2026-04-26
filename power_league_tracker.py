#!/usr/bin/env python3
"""
NCVA Power League Tracker
Fetches live data from Google Sheets and USAV SharePoint,
cross-references bids, and generates an interactive HTML dashboard.
"""

import re
import io
import sys
import csv
import tempfile
import os
import json
import threading
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler

try:
    import requests
except ImportError:
    print("Installing requests...")
    os.system(f"{sys.executable} -m pip install requests")
    import requests

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────

SHEET_ID = "1_Xog0a8Lqf6COYTfp0B8575teSsfQoy5"

AGE_GROUPS = {
    "11":  "239795665",
    "12":  "539092209",
    "13":  "476909113",
    "14":  "2086291804",
    "15":  "486749021",
    "16":  "2061013532",
    "17":  "1333744874",
    "18":  "1333744874",
}

SHAREPOINT_FILES = [
    {
        "url": "https://usavolley.sharepoint.com/:x:/g/IQBVtV5lS9J6TpQDoXuti89uAcKBrSWbQUPJTtMtvBVJS2k?e=D5MudQ",
        "unique_id": "655eb555-d24b-4e7a-9403-a17bad8bcf6e",
        "sheets": [
            "11 National", "11 American",
            "12 National", "12 USA", "12 American",
            "13 Open", "13 National", "13 USA", "13 Liberty", "13 American",
        ],
        "label": "11s-13s",
    },
    {
        "url": "https://usavolley.sharepoint.com/:x:/g/IQCpQtJYjEWoRpSAIYkZnZDbAbko4A5DUvr67W5YWFTIr3g?e=1sTURt",
        "unique_id": "58d242a9-458c-46a8-9480-2189199d90db",
        "sheets": [
            "14 Open", "14 National", "14 USA", "14 Liberty", "14 American", "14 Freedom",
            "15 Open", "15 National", "15 USA", "15 Liberty", "15 American", "15 Freedom",
            "16 Open", "16 National", "16 USA", "16 Liberty", "16 American", "16 Freedom",
            "17 Open", "17 National", "17 USA", "17 Liberty", "17 American", "17 Freedom",
        ],
        "label": "14s-17s",
    },
]

# Official USAV 2025-2026 GJNC Regional Bid Allocations for Northern California
# Source: https://usavolleyball.org/wp-content/uploads/2025/10/2026-Regional-Bid-Allocations.pdf
NORCAL_BID_ALLOCS = {
    "11":    "1 National / 1 American",
    "12":    "1 National / 2 American",
    "13":    "2 National / 2 American",
    "14":    "2 National / 2 American / 2 Freedom",
    "15":    "2 National / 2 American / 2 Freedom",
    "16":  "2 National / 2 American / 2 Freedom",
    "17":  "2 National / 2 American / 2 Freedom",
    "18":  "2 National / 2 American / 2 Freedom",
}

# Manual bid entries for results not yet in the SharePoint files
# Format: team_code -> [(age, bid_type, qualifying_event), ...]
MANUAL_BIDS = {}

TEAM_CODE_RE = re.compile(r"^G\d{2}[A-Z0-9]{3,8}[A-Z]{2}$")

RENO_CSV = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Reno_team_registrations.csv")
OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "NorCal_Power_League_Dashboard.html")


# ─────────────────────────────────────────────────────────────────────────────
# DATA FETCHING — Google Sheets
# ─────────────────────────────────────────────────────────────────────────────

def fetch_age_group(age_label: str, gid: str, session: requests.Session) -> list[dict]:
    """Download CSV for one age group and return list of team dicts.

    CSV column layout (most age groups):
      0: Division  1: Div Place  2: Overall Rank  3: (empty)  4: (div-rank?)
      5: Team Name  6: Team Code  7: PLQ Place  8-10: L1  11-13: L2  14-16: L3
      17-19: Region  20: Season Total  21: Bid note
    The 13s sheet is an outlier: col 7 holds the cumulative total (not PLQ
    place) and col 20 ("Total") is empty.  The parser handles both layouts.
    """
    url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid={gid}"
    print(f"  Fetching {age_label} age group ...", end=" ", flush=True)

    # Cache CSV to avoid re-downloading when multiple age groups share a sheet
    if not hasattr(fetch_age_group, "_cache"):
        fetch_age_group._cache = {}
    if gid in fetch_age_group._cache:
        rows = fetch_age_group._cache[gid]
    else:
        try:
            resp = session.get(url, allow_redirects=True, timeout=30)
            resp.raise_for_status()
        except Exception as e:
            print(f"FAILED ({e})")
            return [], ""
        text = resp.content.decode("utf-8", errors="replace")
        rows = list(csv.reader(io.StringIO(text)))
        fetch_age_group._cache[gid] = rows

    # Extract bid allocation from header rows (e.g., "2 National / 2 American / 2 Freedom")
    bid_alloc = ""
    for row in rows[:5]:
        line = " ".join(c for c in row if c).strip()
        m = re.search(r"(\d+\s+(?:National|Open)[\s/\d\w]*)", line, re.I)
        if m:
            bid_alloc = m.group(1).strip()
            break

    # Detect the code column index from the first matching data row
    # (always index 6 in practice, but we confirm dynamically)
    code_col = None
    for row in rows:
        for i, cell in enumerate(row):
            if TEAM_CODE_RE.match(cell.strip()):
                code_col = i
                break
        if code_col is not None:
            break

    if code_col is None:
        print("found 0 teams (no team codes detected)")
        return [], bid_alloc

    # Derive fixed column positions relative to code column
    # Based on confirmed CSV layout:
    #   code_col = 6, team_name = 5, overall = 2, div_place = 1, division = 0
    col_team_name = code_col - 1   # 5
    col_overall   = code_col - 4   # 2
    col_div_place = code_col - 5   # 1
    col_division  = code_col - 6   # 0

    # Find the "Total" column from the header row.  Most sheets put the season
    # total in this column (around index 20).  The 13s sheet is an outlier: its
    # "Total" column is empty and the cumulative total lives in code_col+1.
    col_total = None
    for row in rows[:6]:
        for i, cell in enumerate(row):
            if cell.strip().lower() == "total":
                col_total = i
                break
        if col_total is not None:
            break

    # Filter by age prefix when multiple age groups share the same sheet (17/18)
    age_prefix = f"G{age_label}" if age_label.isdigit() else None

    teams = []
    for row in rows:
        code = safe_get(row, code_col).strip()
        if not TEAM_CODE_RE.match(code):
            continue
        if age_prefix and not code.startswith(age_prefix):
            continue

        team_name    = safe_get(row, col_team_name).strip()
        overall_rank = safe_get(row, col_overall).strip()
        div_place    = safe_get(row, col_div_place).strip()
        division     = safe_get(row, col_division).strip()

        if not team_name or not re.search(r"[A-Za-z]", team_name):
            continue
        if not re.match(r"^\d+$", overall_rank):
            overall_rank = ""

        # Total points: prefer the header-identified "Total" column, fall back
        # to code_col+1 (where the 13s sheet stores cumulative points).
        total_pts = ""
        if col_total is not None:
            total_pts = safe_get(row, col_total).strip()
        if not re.match(r"^\d+(\.\d+)?$", total_pts):
            # Fallback: code_col+1 may hold the cumulative total (13s layout)
            fallback = safe_get(row, code_col + 1).strip()
            if re.match(r"^\d+(\.\d+)?$", fallback):
                total_pts = fallback

        bid_note = ""
        for j in range(code_col + 1, len(row)):
            val = row[j].strip()
            if val and re.search(r"PNQ|bid|open|national|USA|liberty|american|freedom", val, re.I):
                bid_note = val

        # Strip trailing ".00" for cleaner display, keep decimals otherwise
        if total_pts.endswith(".00"):
            total_pts = total_pts[:-3]

        teams.append({
            "age": age_label,
            "division": division,
            "div_place": div_place,
            "overall_rank": overall_rank,
            "team_name": team_name,
            "team_code": code,
            "total_points": total_pts,
            "bid_status": bid_note,
            "bids": [],  # filled in during cross-reference
        })

    print(f"found {len(teams)} teams" + (f" | Bids: {bid_alloc}" if bid_alloc else ""))
    return teams, bid_alloc


def safe_get(lst, idx):
    if 0 <= idx < len(lst):
        return lst[idx]
    return ""


# ─────────────────────────────────────────────────────────────────────────────
# DATA FETCHING — SharePoint
# ─────────────────────────────────────────────────────────────────────────────

def download_sharepoint_xlsx(file_cfg: dict, session: requests.Session) -> bytes | None:
    """Fetch SharePoint sharing page, extract tempauth URL, download xlsx bytes."""
    url = file_cfg["url"]
    unique_id = file_cfg["unique_id"]
    label = file_cfg["label"]
    print(f"  Fetching SharePoint file ({label}) ...", end=" ", flush=True)

    # Full browser-like headers required by SharePoint to avoid connection reset
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "none",
        "Sec-Fetch-User": "?1",
    }

    try:
        resp = session.get(url, headers=headers, allow_redirects=True, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        print(f"FAILED (page fetch: {e})")
        return None

    html = resp.text
    # The HTML contains JSON-encoded URLs where & is \u0026 — decode that
    html_fixed = html.replace("\\u0026", "&").replace("\\u003d", "=")

    download_url = None

    # Strategy 1: look for download.aspx with tempauth token
    m = re.search(
        r'(https?://[^\s"\'<>\\]*download\.aspx[^\s"\'<>\\]*tempauth=[^\s"\'<>\\]+)',
        html_fixed,
    )
    if m:
        download_url = m.group(1).replace("&amp;", "&")

    # Strategy 2: any URL containing the unique ID
    if not download_url:
        m = re.search(
            r'(https?://[^\s"\'<>\\]*' + re.escape(unique_id) + r'[^\s"\'<>\\]*)',
            html_fixed,
            re.I,
        )
        if m:
            download_url = m.group(1).replace("&amp;", "&")

    if not download_url:
        # Strategy 3: direct download via sharing token append ?download=1
        print("no tempauth found, trying direct download ...")
        direct = build_direct_download_url(url, unique_id, session, headers)
        if direct:
            return direct
        print(f"  FAILED (could not extract download URL for {label})")
        return None
    try:
        dl_resp = session.get(download_url, headers=headers, allow_redirects=True, timeout=60)
        dl_resp.raise_for_status()
        content = dl_resp.content
        # Validate it's actually an xlsx
        if content[:4] == b"PK\x03\x04":
            print(f"downloaded {len(content):,} bytes")
            return content
        else:
            print(f"FAILED (not a valid xlsx, got: {content[:50]})")
            return None
    except Exception as e:
        print(f"FAILED (download: {e})")
        return None


def build_direct_download_url(sharing_url: str, unique_id: str, session: requests.Session, headers: dict) -> bytes | None:
    """Fallback: try to build a direct download URL from the sharing token."""
    # Extract the sharing token from the URL
    # Format: https://usavolley.sharepoint.com/:x:/g/IQ<token>
    m = re.search(r"sharepoint\.com/:[^/]+/g/([^?&]+)", sharing_url)
    if not m:
        return None

    token = m.group(1)
    # Try the download endpoint
    download_url = f"https://usavolley.sharepoint.com/:x:/g/{token}?download=1"
    try:
        resp = session.get(download_url, headers=headers, allow_redirects=True, timeout=60)
        if resp.status_code == 200 and resp.content[:4] == b"PK\x03\x04":
            print(f"direct download succeeded, {len(resp.content):,} bytes")
            return resp.content
    except Exception:
        pass
    return None


def parse_bid_xlsx(xlsx_bytes: bytes, sheet_names: list[str]) -> tuple[dict[str, list[tuple]], dict[str, list[tuple]]]:
    """
    Parse xlsx bytes. Returns:
      bid_map: {team_code: [(age, bid_type, event_qualified), ...]}
      region_extras: {region: [(age, bid_type, event_qualified), ...]}
        Captures BACK TO REGION rows — bids that trickled down to a region
        because the original qualifier-winner already had a higher bid.
    """
    bid_map = {}
    region_extras = {}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception as e:
        print(f"  WARNING: Could not parse xlsx: {e}")
        return bid_map, region_extras

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue

        # Parse "13 National" → age=13, bid_type=National
        parts = sheet_name.strip().split(None, 1)
        if len(parts) < 2:
            continue
        age_str, bid_type = parts[0], parts[1]

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        for row in rows[3:]:  # Skip first 3 header rows
            if not row or len(row) < 4:
                continue
            # Column D (index 3) = team code, E = region, F = event qualified
            code_val = row[3]
            region = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            event_q = str(row[5]).strip() if len(row) > 5 and row[5] else ""

            code = str(code_val).strip() if code_val is not None else ""
            if TEAM_CODE_RE.match(code):
                bid_map.setdefault(code, []).append((age_str, bid_type, event_q))
            elif region and "BACK TO REGION" in event_q.upper():
                # Trickle-down bid returning to a region's pool
                region_extras.setdefault(region, []).append((age_str, bid_type, event_q))

    wb.close()
    return bid_map, region_extras


# ─────────────────────────────────────────────────────────────────────────────
# CROSS-REFERENCE
# ─────────────────────────────────────────────────────────────────────────────

def cross_reference(all_teams: list[dict], bid_map: dict[str, list[tuple]]) -> list[dict]:
    """Attach bid info to each team."""
    for team in all_teams:
        code = team["team_code"]
        if code in bid_map:
            team["bids"] = bid_map[code]
        else:
            team["bids"] = []
    return all_teams


# ─────────────────────────────────────────────────────────────────────────────
# RENO REGISTRATION DATA
# ─────────────────────────────────────────────────────────────────────────────

def load_reno_registrations() -> dict[str, str]:
    """Load Reno Far Western Qualifier registrations. Returns {team_code: division}."""
    if not os.path.exists(RENO_CSV):
        print("  Reno registrations file not found, skipping.")
        return {}
    reno = {}
    with open(RENO_CSV, encoding="latin-1") as f:
        for row in csv.reader(f):
            if len(row) >= 5:
                code = row[2].strip()
                if TEAM_CODE_RE.match(code):
                    reno[code] = row[4].strip()
    print(f"  Loaded {len(reno)} Reno registrations")
    return reno


# ─────────────────────────────────────────────────────────────────────────────
# HTML GENERATION
# ─────────────────────────────────────────────────────────────────────────────

BID_COLORS = {
    "Open":     ("#FFD700", "#7a6000"),   # gold bg, dark text
    "National": ("#1a56db", "#e8f0fe"),   # blue bg, light text
    "USA":      ("#c81e1e", "#fff0f0"),   # red bg, near-white text
    "Liberty":  ("#057a55", "#e6f7ee"),   # green bg
    "American": ("#6b7280", "#f3f4f6"),   # silver/gray bg
    "Freedom":  ("#7e3af2", "#f5f3ff"),   # purple bg
}


def bid_badge_html(bids: list[tuple]) -> str:
    if not bids:
        return ""
    parts = []
    for (age, bid_type, event_q) in bids:
        bg, fg = BID_COLORS.get(bid_type, ("#374151", "#f9fafb"))
        tooltip = f"{age} {bid_type}" + (f" — {event_q}" if event_q else "")
        parts.append(
            f'<span class="bid-badge" style="background:{bg};color:{fg}" title="{tooltip}">'
            f"{bid_type}</span>"
        )
    return " ".join(parts)


BID_TIER_ORDER = ["Open", "National", "USA", "Liberty", "American", "Freedom"]


def generate_html(all_teams: list[dict], fetch_date: str, sharepoint_ok: bool, reno_map: dict = None, nc_trickle_by_age: dict = None) -> str:
    if reno_map is None:
        reno_map = {}
    if nc_trickle_by_age is None:
        nc_trickle_by_age = {}
    groups = list(AGE_GROUPS.keys())
    tabs_html = []
    panels_html = []

    for i, age in enumerate(groups):
        teams = [t for t in all_teams if t["age"] == age]
        active_tab = "active" if i == 0 else ""
        active_panel = "block" if i == 0 else "none"
        tab_id = f"tab-{age.replace('/', '_')}"

        tabs_html.append(
            f'<button class="tab-btn {active_tab}" onclick="switchTab(\'{tab_id}\')" id="btn-{tab_id}">'
            f"{age}s</button>"
        )

        bid_count = sum(1 for t in teams if t["bids"])

        # Find NorCal VBC club teams (team name contains "NorCal")
        norcal_club_teams = []
        # Find teams with bids
        nc_bid_teams = []
        if teams:
            sorted_teams = sorted(teams, key=lambda t: float(t["total_points"]) if t["total_points"] else 0, reverse=True)
            for idx, t in enumerate(sorted_teams, 1):
                if re.search(r'\bNorCal\b', t["team_name"], re.I):
                    pts_str = f" ({t['total_points']} pts)" if t["total_points"] else ""
                    bid_info = ""
                    if t["bids"]:
                        bid_info = " - " + "/".join(bt for _, bt, _ in t["bids"]) + " Bid"
                    norcal_club_teams.append(f"#{idx} {t['team_name']}{pts_str}{bid_info}")
                if t["bids"]:
                    bid_types = "/".join(bt for _, bt, _ in t["bids"])
                    nc_bid_teams.append(f"#{idx} {t['team_name']} ({bid_types})")

        # NorCal club positions box
        norcal_club_html = ""
        if norcal_club_teams:
            norcal_club_html = (
                f'<div class="norcal-club-summary">'
                f'<span class="norcal-club-label">NorCal VBC Positions:</span> '
                + " &nbsp;|&nbsp; ".join(norcal_club_teams)
                + "</div>"
            )

        # Bid holders box
        bid_positions = ""
        if nc_bid_teams:
            bid_positions = (
                f'<div class="nc-bids-summary">'
                f'<span class="nc-bids-label">Teams with Bids:</span> '
                + " &nbsp;|&nbsp; ".join(nc_bid_teams)
                + "</div>"
            )

        # Region bid allocation
        bid_alloc_html = ""
        alloc_str = NORCAL_BID_ALLOCS.get(age, "")
        if alloc_str:
            bid_alloc_html = (
                f'<span class="sep">•</span>'
                f'<span class="stat alloc-stat">Region Bids: {alloc_str}</span>'
            )

        # Trickle-down bids returned to NorCal (BACK TO REGION entries)
        trickle_list = nc_trickle_by_age.get(age, [])
        trickle_alloc_html = ""
        if trickle_list:
            # Aggregate by bid type for compact summary
            tcount = {}
            for bt, _ev in trickle_list:
                tcount[bt] = tcount.get(bt, 0) + 1
            tier_index = {t: i for i, t in enumerate(BID_TIER_ORDER)}
            summary = " / ".join(
                f"{cnt} {bt}" for bt, cnt in sorted(tcount.items(), key=lambda kv: tier_index.get(kv[0], 99))
            )
            sources = ", ".join(f"{bt} from {ev}" for bt, ev in trickle_list)
            trickle_alloc_html = (
                f'<span class="sep">•</span>'
                f'<span class="stat trickle-stat" title="{sources}">'
                f'Trickle-Down: {summary}</span>'
            )

        # ── Bid Projection ────────────────────────────────────────────
        # Parse bid allocation counts from string like "2 National / 2 American / 2 Freedom"
        bid_projection_html = ""
        if (alloc_str or trickle_list) and teams:
            bid_type_counts = []
            for part in alloc_str.split("/") if alloc_str else []:
                m = re.match(r"\s*(\d+)\s+(\w+)", part.strip())
                if m:
                    bid_type_counts.append((m.group(2), int(m.group(1))))

            # Merge in trickle-down bids, then re-sort by tier (Open > National > USA > Liberty > American > Freedom)
            if trickle_list:
                merged = {}
                for bt, cnt in bid_type_counts:
                    merged[bt] = merged.get(bt, 0) + cnt
                for bt, _ev in trickle_list:
                    merged[bt] = merged.get(bt, 0) + 1
                tier_index = {t: i for i, t in enumerate(BID_TIER_ORDER)}
                bid_type_counts = sorted(
                    merged.items(),
                    key=lambda kv: tier_index.get(kv[0], 99),
                )

            # Walk sorted teams, skip those with existing bids, assign region bids
            projected = []  # (bid_type, rank, team)
            skipped = []    # (rank, team_name, existing_bid_types)
            remaining_bids = list(bid_type_counts)  # [(type, count), ...]
            bid_idx = 0  # which bid type we're filling

            for idx, t in enumerate(sorted_teams, 1):
                if bid_idx >= len(remaining_bids):
                    break
                if t["bids"]:
                    bid_strs = "/".join(bt for _, bt, _ in t["bids"])
                    skipped.append((idx, t["team_name"], bid_strs))
                    continue
                bid_type, count = remaining_bids[bid_idx]
                projected.append((bid_type, idx, t))
                remaining_bids[bid_idx] = (bid_type, count - 1)
                if remaining_bids[bid_idx][1] <= 0:
                    bid_idx += 1

            if projected:
                proj_lines = []
                for bt, rank, t in projected:
                    pts = t["total_points"] or "?"
                    proj_lines.append(
                        f'<div class="proj-line">'
                        f'<span class="proj-bid-type">{bt}</span>'
                        f'<span class="proj-arrow">→</span>'
                        f'<span class="proj-team">#{rank} {t["team_name"]} ({pts} pts)</span>'
                        f'</div>'
                    )

                last_rank = projected[-1][1]
                last_pts = float(projected[-1][2]["total_points"]) if projected[-1][2]["total_points"] else 0
                skipped_html = ""
                if skipped:
                    skip_items = [f'{name} ({bids})' for _, name, bids in skipped if _ <= last_rank + 3]
                    if skip_items:
                        skipped_html = (
                            f'<div class="proj-skipped">'
                            f'Skipped (already have bids): {", ".join(skip_items)}'
                            f'</div>'
                        )

                # ── Bid Race Table: show bubble teams with point gaps ──
                # Show teams near the bid line: last 2 bid recipients + next ~6 eligible teams
                bid_race_rows = []
                eligible_above = []  # bid recipients near the line
                eligible_below = []  # teams just missing out
                for idx, t in enumerate(sorted_teams, 1):
                    if t["bids"]:
                        continue
                    t_pts = float(t["total_points"]) if t["total_points"] else 0
                    is_projected = any(r == idx for _, r, _ in projected)
                    if is_projected:
                        eligible_above.append((idx, t, t_pts))
                    else:
                        eligible_below.append((idx, t, t_pts))
                        if len(eligible_below) >= 6:
                            break

                # Build the race table: last 2 bid recipients + bubble teams
                show_above = eligible_above[-2:] if len(eligible_above) >= 2 else eligible_above
                for idx, t, t_pts in show_above:
                    gap = t_pts - last_pts
                    gap_str = f"+{gap:.1f}" if gap > 0 else "BID LINE"
                    bid_race_rows.append(
                        f'<tr class="race-in">'
                        f'<td>#{idx}</td>'
                        f'<td>{t["team_name"]}</td>'
                        f'<td class="num">{t["total_points"]}</td>'
                        f'<td class="race-gap-in">{gap_str}</td>'
                        f'<td class="race-need">Projected bid</td>'
                        f'</tr>'
                    )

                for idx, t, t_pts in eligible_below:
                    gap = last_pts - t_pts
                    # Each regional place ~3 pts apart
                    places_needed = max(1, round(gap / 3))
                    bid_race_rows.append(
                        f'<tr class="race-out">'
                        f'<td>#{idx}</td>'
                        f'<td>{t["team_name"]}</td>'
                        f'<td class="num">{t["total_points"]}</td>'
                        f'<td class="race-gap-out">-{gap:.1f}</td>'
                        f'<td class="race-need">Need ~{places_needed} places higher at Regionals</td>'
                        f'</tr>'
                    )

                race_html = ""
                if bid_race_rows:
                    race_html = (
                        f'<div class="bid-race">'
                        f'<div class="race-header">Bid Race — Point Gaps to Bid Line ({last_pts} pts)</div>'
                        f'<table class="race-table">'
                        f'<thead><tr><th>Rank</th><th>Team</th><th>Points</th><th>Gap</th><th>What it Takes</th></tr></thead>'
                        f'<tbody>{"".join(bid_race_rows)}</tbody>'
                        f'</table>'
                        f'<div class="race-note">Each place at Regionals ≈ 3 pts difference. '
                        f'Gold 1st = ~900 pts, Silver 1st = ~861 pts. '
                        f'To close a 15-pt gap, finish ~5 places higher than the bid-line team at Regionals.</div>'
                        f'</div>'
                    )

                bid_projection_html = (
                    f'<div class="bid-projection">'
                    f'<div class="proj-header">Region Bid Projection <span class="proj-sub">(based on current standings, before Region Championship)</span></div>'
                    + "".join(proj_lines)
                    + f'<div class="proj-bidline">Bid line: approximately rank #{last_rank} ({last_pts} pts)</div>'
                    + skipped_html
                    + race_html
                    + f'<div class="proj-caveat">Region Championship points (~600-900) will shift final standings. '
                    + f'See estimated region points below.</div>'
                    + f'</div>'
                )

        # ── Region Points Reference (collapsible) ────────────────────
        region_ref_html = ""
        if teams:
            # Based on 2024-2025 structure: Gold 1st=900, -6 for 2nd, then -3 per place
            # Each division has 12 teams, next division starts 3 below previous last
            div_names = ["Gold", "Silver", "Bronze", "Aqua", "Blue", "Copper",
                         "Dusk", "Evergreen", "Fuchsia", "Green"]
            ref_rows = []
            base = 900
            for d in div_names:
                first = base
                last = base - 3 * 12 + 3  # 12th place: base - 33
                # Actually: 1st=base, 2nd=base-6, Nth(N>=2)=base-3*N, 12th=base-36
                last = base - 36
                ref_rows.append(f'<span class="ref-div">{d}:</span> {first} → {last}')
                base = last - 3  # next division starts 3 below
            region_ref_html = (
                f'<details class="region-ref">'
                f'<summary>Est. Region Championship Points (based on 2024-2025)</summary>'
                f'<div class="ref-grid">{"<br>".join(ref_rows)}</div>'
                f'<div class="ref-note">12 teams per division. 1st place in each division gets the higher value.</div>'
                f'</details>'
            )

        summary_html = (
            f'<div class="summary">'
            f'<span class="stat">{len(teams)} teams</span>'
            f'<span class="sep">•</span>'
            f'<span class="stat bid-stat">{bid_count} with bids</span>'
            + bid_alloc_html
            + trickle_alloc_html
            + "</div>"
            + norcal_club_html
            + bid_positions
            + bid_projection_html
            + region_ref_html
        )

        rows_html = []
        for rank_idx, t in enumerate(sorted(teams, key=lambda x: float(x["total_points"]) if x["total_points"] else 0, reverse=True), 1):
            row_classes = []
            if t["bids"]:
                row_classes.append("has-bid")
            if re.search(r'\bNorCal\b', t["team_name"], re.I):
                row_classes.append("norcal-club")
            bid_types = ", ".join(bt for _, bt, _ in t["bids"])
            bid_events = "; ".join(ev for _, _, ev in t["bids"] if ev)
            fwq_div = reno_map.get(t["team_code"], "")
            fwq_html = f'<span class="fwq-badge">{fwq_div}</span>' if fwq_div else ""
            rows_html.append(
                f'<tr class="{" ".join(row_classes)}">'
                f'<td data-val="{rank_idx}">{rank_idx}</td>'
                f'<td data-val="{t["division"]}">{t["division"]}</td>'
                f'<td data-val="{t["team_name"]}">{t["team_name"]}</td>'
                f'<td data-val="{t["team_code"]}" class="mono">{t["team_code"]}</td>'
                f'<td data-val="{t["total_points"] or 0}" class="num">{t["total_points"]}</td>'
                f'<td data-val="{bid_types}">{bid_badge_html(t["bids"])}</td>'
                f'<td data-val="{bid_events}">{bid_events}</td>'
                f'<td data-val="{fwq_div}">{fwq_html}</td>'
                "</tr>"
            )

        table_html = f"""
        <table class="data-table" id="tbl-{tab_id}">
          <thead>
            <tr>
              <th onclick="sortTable('tbl-{tab_id}',0)" title="Sort">Rank <span class="sort-icon">⇅</span></th>
              <th onclick="sortTable('tbl-{tab_id}',1)" title="Sort">Division <span class="sort-icon">⇅</span></th>
              <th onclick="sortTable('tbl-{tab_id}',2)" title="Sort">Team Name <span class="sort-icon">⇅</span></th>
              <th onclick="sortTable('tbl-{tab_id}',3)" title="Sort">Team Code <span class="sort-icon">⇅</span></th>
              <th onclick="sortTable('tbl-{tab_id}',4)" title="Sort">Points <span class="sort-icon">⇅</span></th>
              <th onclick="sortTable('tbl-{tab_id}',5)" title="Sort">Bid Type <span class="sort-icon">⇅</span></th>
              <th onclick="sortTable('tbl-{tab_id}',6)" title="Sort">Qualifying Event <span class="sort-icon">⇅</span></th>
              <th onclick="sortTable('tbl-{tab_id}',7)" title="Sort">FWQ Reno <span class="sort-icon">⇅</span></th>
            </tr>
          </thead>
          <tbody>
            {"".join(rows_html) if rows_html else '<tr><td colspan="8" style="text-align:center;padding:2rem;color:#9ca3af;">No data available for this age group</td></tr>'}
          </tbody>
        </table>"""

        panels_html.append(
            f'<div class="tab-panel" id="{tab_id}" style="display:{active_panel}">'
            + summary_html
            + table_html
            + "</div>"
        )

    sharepoint_warn = ""
    if not sharepoint_ok:
        sharepoint_warn = (
            '<div class="warn-banner">'
            "⚠️  Could not fetch USAV bid data from SharePoint. "
            "Bid columns will be empty. NCVA Power League rankings are still shown."
            "</div>"
        )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>NCVA Power League Rankings & Bid Tracker 2025-2026</title>
<style>
  :root {{
    --bg: #0f172a;
    --surface: #1e293b;
    --surface2: #273549;
    --border: #334155;
    --text: #f1f5f9;
    --muted: #94a3b8;
    --accent: #38bdf8;
    --green: #4ade80;
    --green-bg: #14532d;
    --green-dim: #166534;
    --radius: 8px;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    background: var(--bg);
    color: var(--text);
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
    min-height: 100vh;
  }}
  header {{
    background: linear-gradient(135deg, #0f172a 0%, #1e3a5f 100%);
    border-bottom: 1px solid var(--border);
    padding: 1.5rem 2rem;
  }}
  .header-row {{
    display: flex;
    justify-content: space-between;
    align-items: center;
    gap: 1rem;
  }}
  header h1 {{
    font-size: 1.6rem;
    font-weight: 700;
    color: var(--accent);
    letter-spacing: -0.02em;
  }}
  header p {{
    color: var(--muted);
    font-size: 0.85rem;
    margin-top: 0.3rem;
  }}
  .refresh-btn {{
    background: var(--accent);
    color: #0f172a;
    border: none;
    padding: 0.6rem 1.25rem;
    border-radius: var(--radius);
    font-size: 0.9rem;
    font-weight: 600;
    cursor: pointer;
    display: flex;
    align-items: center;
    gap: 0.4rem;
    white-space: nowrap;
    transition: background 0.15s, transform 0.1s;
  }}
  .refresh-btn:hover {{ background: #7dd3fc; }}
  .refresh-btn:active {{ transform: scale(0.97); }}
  .refresh-btn:disabled {{ opacity: 0.6; cursor: not-allowed; }}
  .refresh-icon {{ font-size: 1.1rem; display: inline-block; }}
  .refresh-btn.spinning .refresh-icon {{
    animation: spin 1s linear infinite;
  }}
  @keyframes spin {{ from {{ transform: rotate(0deg); }} to {{ transform: rotate(360deg); }} }}
  .refresh-status {{
    color: var(--muted);
    font-size: 0.8rem;
    margin-top: 0.25rem;
  }}
  .refresh-status.error {{ color: #f87171; }}
  .refresh-status.success {{ color: var(--green); }}
  .norcal-club-summary {{
    background: #1e1b4b;
    border: 1px solid #4338ca;
    border-radius: var(--radius);
    padding: 0.6rem 1rem;
    margin-top: 0.5rem;
    font-size: 0.85rem;
    color: #a5b4fc;
    line-height: 1.6;
  }}
  .norcal-club-label {{
    font-weight: 700;
    color: #c7d2fe;
  }}
  .nc-bids-summary {{
    background: var(--green-bg);
    border: 1px solid var(--green-dim);
    border-radius: var(--radius);
    padding: 0.6rem 1rem;
    margin-top: 0.5rem;
    font-size: 0.85rem;
    color: var(--green);
    line-height: 1.6;
  }}
  .nc-bids-label {{
    font-weight: 700;
    color: #86efac;
  }}
  .alloc-stat {{
    color: #fbbf24 !important;
    font-weight: 600;
  }}
  tr.norcal-club {{
    background: #1e1b4b !important;
    border-left: 3px solid #818cf8;
  }}
  tr.norcal-club td {{ color: #c7d2fe; }}
  .bid-projection {{
    background: #451a03;
    border: 1px solid #b45309;
    border-radius: var(--radius);
    padding: 0.75rem 1rem;
    margin-top: 0.5rem;
    font-size: 0.85rem;
    color: #fcd34d;
  }}
  .proj-header {{
    font-weight: 700;
    color: #fbbf24;
    margin-bottom: 0.5rem;
    font-size: 0.95rem;
  }}
  .proj-sub {{ font-weight: 400; color: #d97706; font-size: 0.8rem; }}
  .proj-line {{
    display: flex;
    align-items: center;
    gap: 0.5rem;
    padding: 0.2rem 0;
  }}
  .proj-bid-type {{
    display: inline-block;
    min-width: 5.5rem;
    padding: 0.1rem 0.5rem;
    border-radius: 999px;
    font-size: 0.75rem;
    font-weight: 700;
    background: #78350f;
    color: #fde68a;
    text-align: center;
  }}
  .proj-arrow {{ color: #92400e; font-weight: 700; }}
  .proj-team {{ color: #fef3c7; }}
  .proj-bidline {{
    margin-top: 0.5rem;
    padding-top: 0.4rem;
    border-top: 1px solid #78350f;
    color: #f59e0b;
    font-weight: 600;
  }}
  .proj-skipped {{
    margin-top: 0.3rem;
    color: #d97706;
    font-size: 0.8rem;
  }}
  .proj-caveat {{
    margin-top: 0.4rem;
    color: #92400e;
    font-size: 0.78rem;
    font-style: italic;
  }}
  .bid-race {{
    margin-top: 0.6rem;
    padding-top: 0.6rem;
    border-top: 1px solid #78350f;
  }}
  .race-header {{
    font-weight: 700;
    color: #f59e0b;
    margin-bottom: 0.4rem;
    font-size: 0.9rem;
  }}
  .race-table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 0.82rem;
  }}
  .race-table th {{
    text-align: left;
    padding: 0.3rem 0.5rem;
    color: #d97706;
    border-bottom: 1px solid #78350f;
    font-weight: 600;
  }}
  .race-table td {{
    padding: 0.3rem 0.5rem;
    border-bottom: 1px solid rgba(120,53,15,0.3);
  }}
  tr.race-in td {{ color: #86efac; }}
  tr.race-out td {{ color: #fcd34d; }}
  .race-gap-in {{ font-weight: 700; color: #4ade80 !important; }}
  .race-gap-out {{ font-weight: 700; color: #fb923c !important; }}
  .race-need {{ font-style: italic; font-size: 0.78rem; }}
  tr.race-in .race-need {{ color: #86efac; }}
  tr.race-out .race-need {{ color: #fbbf24; }}
  .race-note {{
    margin-top: 0.4rem;
    color: #92400e;
    font-size: 0.75rem;
    line-height: 1.5;
  }}
  .region-ref {{
    margin-top: 0.5rem;
    background: var(--surface2);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    font-size: 0.8rem;
  }}
  .region-ref summary {{
    padding: 0.5rem 1rem;
    cursor: pointer;
    color: var(--muted);
    font-weight: 600;
  }}
  .region-ref summary:hover {{ color: var(--text); }}
  .ref-grid {{
    padding: 0.5rem 1rem;
    color: var(--muted);
    line-height: 1.6;
  }}
  .ref-div {{ color: var(--accent); font-weight: 600; display: inline-block; min-width: 5rem; }}
  .ref-note {{ padding: 0.3rem 1rem 0.6rem; color: #64748b; font-size: 0.75rem; }}
  .fwq-badge {{
    display: inline-block;
    padding: 0.15rem 0.5rem;
    border-radius: 999px;
    font-size: 0.75rem;
    font-weight: 600;
    background: #1e3a5f;
    color: #7dd3fc;
    border: 1px solid #38bdf8;
  }}
  .warn-banner {{
    background: #451a03;
    border: 1px solid #92400e;
    color: #fcd34d;
    padding: 0.75rem 2rem;
    font-size: 0.875rem;
  }}
  .tabs {{
    background: var(--surface);
    border-bottom: 1px solid var(--border);
    padding: 0 2rem;
    display: flex;
    gap: 0.25rem;
    overflow-x: auto;
  }}
  .tab-btn {{
    background: none;
    border: none;
    color: var(--muted);
    padding: 0.875rem 1.25rem;
    cursor: pointer;
    font-size: 0.925rem;
    font-weight: 500;
    border-bottom: 2px solid transparent;
    white-space: nowrap;
    transition: color 0.15s, border-color 0.15s;
  }}
  .tab-btn:hover {{ color: var(--text); }}
  .tab-btn.active {{ color: var(--accent); border-bottom-color: var(--accent); }}
  .tab-panel {{ padding: 1.5rem 2rem; }}
  .summary {{
    display: flex;
    align-items: center;
    gap: 0.5rem;
    margin-bottom: 1rem;
    flex-wrap: wrap;
  }}
  .summary .stat {{
    background: var(--surface2);
    padding: 0.35rem 0.75rem;
    border-radius: 999px;
    font-size: 0.825rem;
    color: var(--muted);
  }}
  .summary .bid-stat {{ color: var(--green); background: var(--green-bg); }}
  .summary .trickle-stat {{ color: #7e3af2; background: #f5f3ff; border: 1px solid #ddd6fe; cursor: help; }}
  .summary .sep {{ color: var(--border); }}
  .data-table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 0.875rem;
    border-radius: var(--radius);
    overflow: hidden;
    border: 1px solid var(--border);
  }}
  .data-table thead {{
    background: var(--surface2);
  }}
  .data-table th {{
    padding: 0.7rem 0.9rem;
    text-align: left;
    font-weight: 600;
    color: var(--muted);
    cursor: pointer;
    user-select: none;
    white-space: nowrap;
    font-size: 0.775rem;
    text-transform: uppercase;
    letter-spacing: 0.04em;
  }}
  .data-table th:hover {{ color: var(--text); }}
  .sort-icon {{ opacity: 0.4; font-size: 0.7rem; }}
  .data-table td {{
    padding: 0.6rem 0.9rem;
    border-top: 1px solid var(--border);
    color: var(--text);
    vertical-align: middle;
  }}
  .data-table tbody tr {{ background: var(--surface); }}
  .data-table tbody tr:nth-child(even) {{ background: var(--surface2); }}
  .data-table tbody tr.has-bid {{ background: var(--green-bg) !important; }}
  .data-table tbody tr.has-bid td {{ color: #bbf7d0; }}
  .data-table tbody tr:hover td {{ background: rgba(56,189,248,0.06); }}
  td.mono {{ font-family: "Consolas", "SF Mono", monospace; font-size: 0.8rem; color: var(--muted); }}
  td.num {{ font-variant-numeric: tabular-nums; }}
  .bid-badge {{
    display: inline-block;
    padding: 0.2rem 0.6rem;
    border-radius: 999px;
    font-size: 0.75rem;
    font-weight: 600;
    cursor: default;
  }}
  @media (max-width: 768px) {{
    header {{ padding: 1rem; }}
    .tab-panel {{ padding: 1rem; }}
    .tabs {{ padding: 0 1rem; }}
    .data-table {{ font-size: 0.8rem; }}
    .data-table td, .data-table th {{ padding: 0.5rem; }}
  }}
</style>
</head>
<body>
<header>
  <div class="header-row">
    <div>
      <h1>NCVA Girls Power League Rankings &amp; Bid Tracker 2025-2026</h1>
      <p>Data fetched on <span id="fetch-date">{fetch_date}</span> &nbsp;•&nbsp; Sources: NCVA Google Sheets, USAV SharePoint bid lists</p>
    </div>
    <div style="text-align:right">
      <button id="refresh-btn" class="refresh-btn" onclick="refreshData()">
        <span class="refresh-icon">&#x21bb;</span> Refresh Data
      </button>
      <div id="refresh-status" class="refresh-status"></div>
    </div>
  </div>
</header>
{sharepoint_warn}
<nav class="tabs">
  {"".join(tabs_html)}
</nav>
<main>
  {"".join(panels_html)}
</main>

<script>
function switchTab(id) {{
  document.querySelectorAll('.tab-panel').forEach(p => p.style.display = 'none');
  document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
  document.getElementById(id).style.display = 'block';
  document.getElementById('btn-' + id).classList.add('active');
}}

const sortState = {{}};
function sortTable(tableId, colIdx) {{
  const table = document.getElementById(tableId);
  const tbody = table.querySelector('tbody');
  const rows = Array.from(tbody.querySelectorAll('tr'));
  const key = tableId + ':' + colIdx;
  const asc = !sortState[key];
  sortState[key] = asc;

  rows.sort((a, b) => {{
    const aCell = a.querySelectorAll('td')[colIdx];
    const bCell = b.querySelectorAll('td')[colIdx];
    if (!aCell || !bCell) return 0;
    let aVal = aCell.getAttribute('data-val') || aCell.textContent.trim();
    let bVal = bCell.getAttribute('data-val') || bCell.textContent.trim();
    const aNum = parseFloat(aVal);
    const bNum = parseFloat(bVal);
    if (!isNaN(aNum) && !isNaN(bNum)) {{
      return asc ? aNum - bNum : bNum - aNum;
    }}
    return asc ? aVal.localeCompare(bVal) : bVal.localeCompare(aVal);
  }});

  rows.forEach(r => tbody.appendChild(r));

  // Update sort icons
  table.querySelectorAll('th .sort-icon').forEach((ic, i) => {{
    ic.textContent = i === colIdx ? (asc ? '↑' : '↓') : '⇅';
    ic.style.opacity = i === colIdx ? '1' : '0.4';
  }});
}}

const GH_REPO = 'musint/ncva-power-league-dashboard';
const GH_WORKFLOW = 'update-dashboard.yml';

function getToken() {{
  let token = localStorage.getItem('gh_pat');
  if (!token) {{
    token = prompt(
      'First-time setup: Enter a GitHub Personal Access Token\\n' +
      'with "Actions: Read & Write" permission for this repo.\\n\\n' +
      'Create one at: https://github.com/settings/tokens?type=beta\\n' +
      'Scope: musint/ncva-power-league-dashboard → Actions: Read and write\\n\\n' +
      'This is saved in your browser only (localStorage).'
    );
    if (token) localStorage.setItem('gh_pat', token.trim());
  }}
  return token;
}}

async function refreshData() {{
  const token = getToken();
  if (!token) return;

  const btn = document.getElementById('refresh-btn');
  const statusEl = document.getElementById('refresh-status');
  btn.disabled = true;
  btn.classList.add('spinning');
  btn.childNodes[btn.childNodes.length - 1].textContent = ' Triggering update...';
  statusEl.textContent = '';
  statusEl.className = 'refresh-status';

  try {{
    // Trigger the workflow
    const dispatchRes = await fetch(
      `https://api.github.com/repos/${{GH_REPO}}/actions/workflows/${{GH_WORKFLOW}}/dispatches`,
      {{
        method: 'POST',
        headers: {{
          'Authorization': `Bearer ${{token}}`,
          'Accept': 'application/vnd.github+json',
        }},
        body: JSON.stringify({{ ref: 'main' }}),
      }}
    );
    if (dispatchRes.status === 401 || dispatchRes.status === 403) {{
      localStorage.removeItem('gh_pat');
      throw new Error('Invalid token. Click Refresh to try again.');
    }}
    if (!dispatchRes.ok) throw new Error('Trigger failed: ' + dispatchRes.status);

    statusEl.textContent = 'Workflow triggered. Waiting for completion...';

    // Poll for workflow completion
    await new Promise(r => setTimeout(r, 5000));
    for (let i = 0; i < 20; i++) {{
      const runsRes = await fetch(
        `https://api.github.com/repos/${{GH_REPO}}/actions/workflows/${{GH_WORKFLOW}}/runs?per_page=1`,
        {{ headers: {{ 'Authorization': `Bearer ${{token}}`, 'Accept': 'application/vnd.github+json' }} }}
      );
      const runs = await runsRes.json();
      const latest = runs.workflow_runs && runs.workflow_runs[0];
      if (latest && latest.status === 'completed') {{
        if (latest.conclusion === 'success') {{
          statusEl.textContent = 'Update complete! Reloading...';
          statusEl.className = 'refresh-status success';
          // Wait for GitHub Pages to deploy (~15s)
          await new Promise(r => setTimeout(r, 15000));
          window.location.reload();
          return;
        }} else {{
          throw new Error('Workflow failed: ' + latest.conclusion);
        }}
      }}
      statusEl.textContent = `Waiting for workflow... (${{(i+1)*5}}s)`;
      await new Promise(r => setTimeout(r, 5000));
    }}
    throw new Error('Workflow timed out after 100s');
  }} catch (err) {{
    statusEl.textContent = err.message;
    statusEl.className = 'refresh-status error';
    btn.disabled = false;
    btn.classList.remove('spinning');
    btn.childNodes[btn.childNodes.length - 1].textContent = ' Refresh Data';
  }}
}}
</script>
</body>
</html>"""

    return html


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

SERVE_PORT = 8765


def fetch_and_generate():
    """Fetch all data, cross-reference, and write the HTML dashboard. Returns summary string."""
    print("=" * 60)
    print("NCVA Power League Tracker")
    print("=" * 60)

    session = requests.Session()
    session.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
    })

    # ── Step 1: Fetch Power League data ──────────────────────────────────────
    print("\n[1/3] Fetching NCVA Power League data from Google Sheets...")
    all_teams = []
    for age, gid in AGE_GROUPS.items():
        teams, _bid_alloc = fetch_age_group(age, gid, session)
        all_teams.extend(teams)
    print(f"  Total teams fetched: {len(all_teams)}")

    # ── Step 2: Fetch SharePoint bid data ────────────────────────────────────
    print("\n[2/3] Fetching USAV bid qualification data from SharePoint...")
    bid_map = {}
    region_extras_all = {}  # region -> [(age, bid_type, event), ...]
    sharepoint_ok = False

    for file_cfg in SHAREPOINT_FILES:
        # Use a fresh session per SharePoint file to avoid stale cookie/connection issues
        sp_session = requests.Session()
        sp_session.headers.update(session.headers)
        xlsx_bytes = download_sharepoint_xlsx(file_cfg, sp_session)
        if xlsx_bytes:
            partial, partial_extras = parse_bid_xlsx(xlsx_bytes, file_cfg["sheets"])
            bid_map.update(partial)
            for region, entries in partial_extras.items():
                region_extras_all.setdefault(region, []).extend(entries)
            sharepoint_ok = True
            print(f"  Parsed {len(partial)} bid-qualified teams from {file_cfg['label']}")
        else:
            print(f"  WARNING: Skipping bid data for {file_cfg['label']}")

    print(f"  Total bid-qualified teams (all regions): {len(bid_map)}")
    nc_bid_codes = {k for k in bid_map if k.endswith("NC")}
    print(f"  NorCal (NC) bid-qualified teams: {len(nc_bid_codes)}")

    # NorCal trickle-down bids returned to the region — group by age
    nc_trickle_by_age = {}  # age -> [(bid_type, event), ...]
    for age, bid_type, event_q in region_extras_all.get("NC", []):
        nc_trickle_by_age.setdefault(age, []).append((bid_type, event_q))
    if nc_trickle_by_age:
        print(f"  NorCal trickle-down bids: {sum(len(v) for v in nc_trickle_by_age.values())}")
        for age in sorted(nc_trickle_by_age):
            for bt, ev in nc_trickle_by_age[age]:
                print(f"    {age}s +1 {bt} (from {ev})")

    # ── Step 2b: Merge manual bid entries ────────────────────────────────────
    for code, entries in MANUAL_BIDS.items():
        if code not in bid_map:
            bid_map[code] = entries
        else:
            # Add manual entries that aren't already present
            existing = {(a, b) for a, b, _ in bid_map[code]}
            for entry in entries:
                if (entry[0], entry[1]) not in existing:
                    bid_map[code].append(entry)
    print(f"  Manual bid entries merged: {len(MANUAL_BIDS)} teams")

    # ── Step 3: Cross-reference ───────────────────────────────────────────────
    print("\n[3/3] Cross-referencing team codes...")
    all_teams = cross_reference(all_teams, bid_map)
    bid_holders = [t for t in all_teams if t["bids"]]
    print(f"  NCVA teams with bids: {len(bid_holders)}")
    for t in bid_holders:
        bid_str = ", ".join(f"{a} {b}" for a, b, _ in t["bids"])
        print(f"    {t['team_code']:20s} {t['team_name']:35s} [{bid_str}]")

    # ── Step 4: Load Reno registrations ─────────────────────────────────────
    print("\n[4/4] Loading Far Western Qualifier registrations...")
    reno_map = load_reno_registrations()

    # ── Generate HTML ─────────────────────────────────────────────────────────
    fetch_date = datetime.now().strftime("%B %d, %Y at %I:%M %p")
    html = generate_html(all_teams, fetch_date, sharepoint_ok, reno_map, nc_trickle_by_age)

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write(html)

    summary = f"{len(all_teams)} total teams | {len(bid_holders)} with bids"
    print(f"\nDashboard written to: {OUTPUT_PATH}")
    print(f"  {summary}")
    return summary


class RefreshHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        if self.path == "/refresh":
            self.send_response(200)
            self.send_header("Access-Control-Allow-Origin", "*")
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            try:
                summary = fetch_and_generate()
                self.wfile.write(json.dumps({"ok": True, "summary": summary}).encode())
            except Exception as e:
                self.wfile.write(json.dumps({"ok": False, "error": str(e)}).encode())
        elif self.path == "/" or self.path == "/dashboard":
            # Serve the dashboard HTML directly
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers()
            with open(OUTPUT_PATH, "r", encoding="utf-8") as f:
                self.wfile.write(f.read().encode())
        else:
            self.send_response(404)
            self.end_headers()

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, OPTIONS")
        self.end_headers()

    def log_message(self, format, *args):
        print(f"  [server] {args[0]}")


def serve():
    """Run as a local web server with a /refresh endpoint."""
    if not os.path.exists(OUTPUT_PATH):
        print("No existing dashboard found. Generating initial dashboard...")
        fetch_and_generate()
    else:
        print(f"Using existing dashboard: {OUTPUT_PATH}")
        print("Click 'Refresh Data' in the browser to fetch fresh data.")

    server = HTTPServer(("127.0.0.1", SERVE_PORT), RefreshHandler)
    print(f"\nServer running at http://localhost:{SERVE_PORT}")
    print(f"  Dashboard: http://localhost:{SERVE_PORT}/dashboard")
    print(f"  Refresh:   http://localhost:{SERVE_PORT}/refresh")
    print(f"\nPress Ctrl+C to stop.\n")

    import webbrowser
    webbrowser.open(f"http://localhost:{SERVE_PORT}/dashboard")

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nServer stopped.")
        server.server_close()


def main():
    if "--serve" in sys.argv:
        serve()
    else:
        fetch_and_generate()
        print("  Open the HTML file in a browser to view the dashboard.")
        print(f"  Or run with --serve for live refresh: python power_league_tracker.py --serve")


if __name__ == "__main__":
    main()
